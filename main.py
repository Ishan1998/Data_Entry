from tkinter import *
from tkinter.ttk import Combobox
import tkinter as tk
from tkinter import messagebox
import openpyxl , xlrd
from openpyxl import Workbook
import pathlib

root=Tk()
root.title("Data Entry Software")
root.geometry('700x400+300+200')
root.resizable(False,False)
root.configure(bg="#040404")

file=pathlib.Path('Data.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']= "Full Name"
    sheet['B1']= "Phone Number"
    sheet['C1']= "Age"
    sheet['D1']= "Gender"
    sheet['E1']= "Address"

    file.save('Data.xlsx')

def submit():
    name=nameValue.get()
    contact=contactValue.get()
    age=ageValue.get()
    gender=gender_combobox.get()
    address=addressEntry.get(1.0,END)

    file=openpyxl.load_workbook('Data.xlsx')
    sheet=file.active
    sheet.cell(column=1,row=sheet.max_row+1,value=name)
    sheet.cell(column=2,row=sheet.max_row,value=contact)
    sheet.cell(column=3,row=sheet.max_row,value=age)
    sheet.cell(column=4,row=sheet.max_row,value=gender)
    sheet.cell(column=5,row=sheet.max_row,value=address)

    file.save(r'Data.xlsx')

def clear():
    nameValue.set('')
    contactValue.set('')
    ageValue.set('')
    addressEntry.delete(1.0,END)

icon_image=PhotoImage(file="logo.png")
root.iconphoto(False,icon_image)

Label(root,text="Please fill out this Entry Form:", bg="#040404", font="Times 24 bold italic",fg="white").place(x=20, y=20)

Label(root,text='Name', font=23, bg="#040404", fg="white").place(x=50,y=100)
Label(root,text='Contact No.', font=23, bg="#040404", fg="white").place(x=50,y=150)
Label(root,text='Age', font=23, bg="#040404", fg="white").place(x=50,y=200)
Label(root,text='Gender', font=23, bg="#040404", fg="white").place(x=370,y=200)
Label(root,text='Address', font=23, bg="#040404", fg="white").place(x=50,y=250)

nameValue = StringVar()
contactValue = StringVar()
ageValue = StringVar()

nameEntry = Entry(root,textvariable=nameValue,width=10,bd=5,font=20)
contactEntry = Entry(root,textvariable=contactValue,width=10,bd=5,font=20)
ageEntry = Entry(root,textvariable=ageValue,width=10,bd=5,font=20)

gender_combobox = Combobox(root,values=['Male', 'Female'], font='Times 17 bold italic', state='r', width=7)
gender_combobox.place(x=450,y=200)
gender_combobox.set('Male')

addressEntry = Text(root,width=50,height=4,bd=2)

nameEntry.place(x=200,y=100)
contactEntry.place(x=200,y=150)
ageEntry.place(x=200,y=200)
addressEntry.place(x=200,y=250)

nameEntry.place(x=200, y=100)
contactEntry.place(x=200, y=150)
ageEntry.place(x=200, y=200)

Button(root,text="Submit", font="Times 13 bold", bg="#6C2896", fg="white",width=15,height=2,command=submit).place(x=50,y=340)
Button(root,text="Clear", font="Times 13 bold", bg="#6C2896", fg="white",width=15,height=2, command=clear).place(x=290,y=340)
Button(root,text="Exit", font="Times 13 bold", bg="#6C2896", fg="white",width=15,height=2, command=lambda:root.destroy()).place(x=520,y=340)

root.mainloop()